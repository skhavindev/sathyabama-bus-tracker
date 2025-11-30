from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_
from typing import List, Optional
from datetime import datetime
import math
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

from ..database import get_db
from ..models.driver import Driver
from ..models.bus_route import BusRoute
from ..models.audit_log import AuditLog
from ..schemas.admin import (
    DriverCreate, DriverUpdate, DriverResponse, DriverListResponse,
    RouteCreate, RouteUpdate, RouteResponse, RouteImportResponse,
    StatisticsResponse, AuditLogResponse, AuditLogListResponse
)
from ..services.auth_service import get_current_admin, hash_password
from ..services.cache_service import CacheService

router = APIRouter(prefix="/api/admin", tags=["Admin"])


# Helper function for audit logging
def create_audit_log(
    db: Session,
    admin_id: int,
    action_type: str,
    entity_type: str,
    entity_id: int,
    changes: dict = None
):
    """Create an audit log entry"""
    audit_log = AuditLog(
        admin_id=admin_id,
        action_type=action_type,
        entity_type=entity_type,
        entity_id=entity_id,
        changes=changes
    )
    db.add(audit_log)
    db.commit()


# ==================== DRIVER MANAGEMENT ====================

@router.get("/drivers", response_model=DriverListResponse)
def list_drivers(
    page: int = 1,
    per_page: int = 10,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """List all drivers with pagination and search"""
    query = db.query(Driver)
    
    # Apply search filter
    if search:
        search_filter = or_(
            Driver.name.ilike(f"%{search}%"),
            Driver.phone.ilike(f"%{search}%"),
            Driver.email.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    # Get total count
    total = query.count()
    
    # Calculate pagination
    pages = math.ceil(total / per_page)
    offset = (page - 1) * per_page
    
    # Get paginated results
    drivers = query.order_by(Driver.created_at.desc()).offset(offset).limit(per_page).all()
    
    return DriverListResponse(
        drivers=drivers,
        total=total,
        page=page,
        pages=pages
    )


@router.post("/drivers", response_model=DriverResponse)
def create_driver(
    driver_data: DriverCreate,
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Create new driver account"""
    # Check if phone already exists
    existing = db.query(Driver).filter(Driver.phone == driver_data.phone).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Phone number already exists"
        )
    
    # Check if email already exists (if provided)
    if driver_data.email:
        existing_email = db.query(Driver).filter(Driver.email == driver_data.email).first()
        if existing_email:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Email already exists"
            )
    
    # Create driver
    new_driver = Driver(
        name=driver_data.name,
        phone=driver_data.phone,
        email=driver_data.email,
        hashed_password=hash_password(driver_data.password),
        is_active=driver_data.is_active,
        is_admin=driver_data.is_admin
    )
    
    db.add(new_driver)
    db.commit()
    db.refresh(new_driver)
    
    # Create audit log
    create_audit_log(
        db, current_admin.driver_id, "CREATE", "driver", new_driver.driver_id,
        {"name": new_driver.name, "phone": new_driver.phone}
    )
    
    return new_driver


@router.get("/drivers/{driver_id}", response_model=DriverResponse)
def get_driver(
    driver_id: int,
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Get driver details"""
    driver = db.query(Driver).filter(Driver.driver_id == driver_id).first()
    if not driver:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Driver not found"
        )
    return driver


@router.put("/drivers/{driver_id}", response_model=DriverResponse)
def update_driver(
    driver_id: int,
    driver_data: DriverUpdate,
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Update driver information"""
    driver = db.query(Driver).filter(Driver.driver_id == driver_id).first()
    if not driver:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Driver not found"
        )
    
    # Store old values for audit
    old_values = {
        "name": driver.name,
        "phone": driver.phone,
        "email": driver.email,
        "is_active": driver.is_active,
        "is_admin": driver.is_admin
    }
    
    # Update fields
    if driver_data.name is not None:
        driver.name = driver_data.name
    if driver_data.phone is not None:
        # Check uniqueness
        existing = db.query(Driver).filter(
            and_(Driver.phone == driver_data.phone, Driver.driver_id != driver_id)
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Phone number already exists"
            )
        driver.phone = driver_data.phone
    if driver_data.email is not None:
        # Check uniqueness
        existing = db.query(Driver).filter(
            and_(Driver.email == driver_data.email, Driver.driver_id != driver_id)
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Email already exists"
            )
        driver.email = driver_data.email
    if driver_data.password is not None:
        driver.hashed_password = hash_password(driver_data.password)
    if driver_data.is_active is not None:
        driver.is_active = driver_data.is_active
    if driver_data.is_admin is not None:
        driver.is_admin = driver_data.is_admin
    
    db.commit()
    db.refresh(driver)
    
    # Create audit log
    new_values = {
        "name": driver.name,
        "phone": driver.phone,
        "email": driver.email,
        "is_active": driver.is_active,
        "is_admin": driver.is_admin
    }
    create_audit_log(
        db, current_admin.driver_id, "UPDATE", "driver", driver_id,
        {"before": old_values, "after": new_values}
    )
    
    return driver


@router.delete("/drivers/{driver_id}")
def delete_driver(
    driver_id: int,
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Delete driver account"""
    driver = db.query(Driver).filter(Driver.driver_id == driver_id).first()
    if not driver:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Driver not found"
        )
    
    # Prevent deleting self
    if driver.driver_id == current_admin.driver_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Cannot delete your own account"
        )
    
    # Store info for audit
    driver_info = {"name": driver.name, "phone": driver.phone}
    
    # Delete driver
    db.delete(driver)
    db.commit()
    
    # Create audit log
    create_audit_log(
        db, current_admin.driver_id, "DELETE", "driver", driver_id,
        driver_info
    )
    
    return {"success": True, "message": "Driver deleted successfully"}


# ==================== BUS ROUTE MANAGEMENT ====================

@router.get("/routes", response_model=List[RouteResponse])
def list_routes(
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """List all bus routes ordered by sl_no"""
    routes = db.query(BusRoute).order_by(BusRoute.sl_no).all()
    return routes


@router.post("/routes", response_model=RouteResponse)
def create_route(
    route_data: RouteCreate,
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Create new bus route and auto-link to driver by phone number"""
    # Check vehicle_no uniqueness
    existing = db.query(BusRoute).filter(BusRoute.vehicle_no == route_data.vehicle_no).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Vehicle number already exists"
        )
    
    # Auto-assign next sl_no
    max_sl_no = db.query(BusRoute).count()
    next_sl_no = max_sl_no + 1
    
    # Try to find driver by phone number to auto-link
    driver_id = route_data.driver_id
    if not driver_id and route_data.phone_number:
        driver = db.query(Driver).filter(Driver.phone == route_data.phone_number).first()
        if driver:
            driver_id = driver.driver_id
    
    # Create route
    new_route = BusRoute(
        sl_no=next_sl_no,
        bus_route=route_data.bus_route,
        route_no=route_data.route_no,
        vehicle_no=route_data.vehicle_no,
        driver_id=driver_id,
        driver_name=route_data.driver_name,
        phone_number=route_data.phone_number,
        is_active=route_data.is_active
    )
    
    db.add(new_route)
    db.commit()
    db.refresh(new_route)
    
    # Create audit log
    create_audit_log(
        db, current_admin.driver_id, "CREATE", "route", new_route.route_id,
        {"route_no": new_route.route_no, "vehicle_no": new_route.vehicle_no, "driver_id": driver_id}
    )
    
    return new_route


@router.put("/routes/{route_id}", response_model=RouteResponse)
def update_route(
    route_id: int,
    route_data: RouteUpdate,
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Update bus route"""
    route = db.query(BusRoute).filter(BusRoute.route_id == route_id).first()
    if not route:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Route not found"
        )
    
    # Store old values
    old_values = {
        "sl_no": route.sl_no,
        "bus_route": route.bus_route,
        "route_no": route.route_no,
        "vehicle_no": route.vehicle_no,
        "driver_name": route.driver_name,
        "phone_number": route.phone_number
    }
    
    # Update fields
    if route_data.sl_no is not None:
        route.sl_no = route_data.sl_no
    if route_data.bus_route is not None:
        route.bus_route = route_data.bus_route
    if route_data.route_no is not None:
        route.route_no = route_data.route_no
    if route_data.vehicle_no is not None:
        # Check uniqueness
        existing = db.query(BusRoute).filter(
            and_(BusRoute.vehicle_no == route_data.vehicle_no, BusRoute.route_id != route_id)
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Vehicle number already exists"
            )
        route.vehicle_no = route_data.vehicle_no
    if route_data.driver_id is not None:
        route.driver_id = route_data.driver_id
    if route_data.driver_name is not None:
        route.driver_name = route_data.driver_name
    if route_data.phone_number is not None:
        route.phone_number = route_data.phone_number
    if route_data.is_active is not None:
        route.is_active = route_data.is_active
    
    db.commit()
    db.refresh(route)
    
    # Create audit log
    new_values = {
        "sl_no": route.sl_no,
        "bus_route": route.bus_route,
        "route_no": route.route_no,
        "vehicle_no": route.vehicle_no,
        "driver_name": route.driver_name,
        "phone_number": route.phone_number
    }
    create_audit_log(
        db, current_admin.driver_id, "UPDATE", "route", route_id,
        {"before": old_values, "after": new_values}
    )
    
    return route


@router.delete("/routes/{route_id}")
def delete_route(
    route_id: int,
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Delete bus route and renumber remaining routes"""
    route = db.query(BusRoute).filter(BusRoute.route_id == route_id).first()
    if not route:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Route not found"
        )
    
    deleted_sl_no = route.sl_no
    route_info = {"route_no": route.route_no, "vehicle_no": route.vehicle_no}
    
    # Delete route
    db.delete(route)
    
    # Renumber remaining routes
    routes_to_renumber = db.query(BusRoute).filter(BusRoute.sl_no > deleted_sl_no).all()
    for r in routes_to_renumber:
        r.sl_no -= 1
    
    db.commit()
    
    # Create audit log
    create_audit_log(
        db, current_admin.driver_id, "DELETE", "route", route_id,
        route_info
    )
    
    return {"success": True, "message": "Route deleted successfully"}


@router.post("/routes/import", response_model=RouteImportResponse)
async def import_routes(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Bulk import routes from Excel/CSV"""
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel (.xlsx, .xls) and CSV files are supported"
        )
    
    imported = 0
    failed = 0
    errors = []
    
    try:
        contents = await file.read()
        
        if file.filename.endswith('.csv'):
            # Parse CSV
            import csv
            lines = contents.decode('utf-8').splitlines()
            reader = csv.DictReader(lines)
            rows = list(reader)
        else:
            # Parse Excel
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        
        # Get current max sl_no
        max_sl_no = db.query(BusRoute).count()
        next_sl_no = max_sl_no + 1
        
        for idx, row in enumerate(rows, start=2):
            try:
                # Validate required fields
                required_fields = ['Bus Route', 'Route No', 'Vehicle No', 'Driver Name', 'Phone Number']
                missing = [f for f in required_fields if not row.get(f)]
                if missing:
                    errors.append({
                        "row": idx,
                        "error": f"Missing fields: {', '.join(missing)}"
                    })
                    failed += 1
                    continue
                
                # Check vehicle_no uniqueness
                vehicle_no = str(row['Vehicle No']).strip()
                existing = db.query(BusRoute).filter(BusRoute.vehicle_no == vehicle_no).first()
                if existing:
                    errors.append({
                        "row": idx,
                        "error": f"Vehicle number {vehicle_no} already exists"
                    })
                    failed += 1
                    continue
                
                # Create route
                new_route = BusRoute(
                    sl_no=next_sl_no,
                    bus_route=str(row['Bus Route']).strip(),
                    route_no=str(row['Route No']).strip(),
                    vehicle_no=vehicle_no,
                    driver_name=str(row['Driver Name']).strip(),
                    phone_number=str(row['Phone Number']).strip(),
                    is_active=True
                )
                
                db.add(new_route)
                next_sl_no += 1
                imported += 1
                
            except Exception as e:
                errors.append({
                    "row": idx,
                    "error": str(e)
                })
                failed += 1
        
        db.commit()
        
        # Create audit log
        create_audit_log(
            db, current_admin.driver_id, "CREATE", "route", 0,
            {"action": "bulk_import", "imported": imported, "failed": failed}
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing file: {str(e)}"
        )
    
    return RouteImportResponse(
        imported=imported,
        failed=failed,
        errors=errors
    )


@router.get("/routes/export")
async def export_routes(
    format: str = "excel",
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Export routes as Excel or PDF"""
    routes = db.query(BusRoute).order_by(BusRoute.sl_no).all()
    
    if format == "excel":
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bus Routes"
        
        # Headers
        headers = ['Sl.No', 'Bus Route', 'Route No', 'Vehicle No', 'Driver Name', 'Phone Number']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for route in routes:
            ws.append([
                route.sl_no,
                route.bus_route,
                route.route_no,
                route.vehicle_no,
                route.driver_name,
                route.phone_number
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"bus_routes_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    elif format == "pdf":
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title = Paragraph("Sathyabama University Bus Routes", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Date
        date_text = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal'])
        elements.append(date_text)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        data = [['Sl.No', 'Bus Route', 'Route No', 'Vehicle No', 'Driver Name', 'Phone Number']]
        for route in routes:
            data.append([
                str(route.sl_no),
                route.bus_route[:50] + '...' if len(route.bus_route) > 50 else route.bus_route,
                route.route_no,
                route.vehicle_no,
                route.driver_name,
                route.phone_number
            ])
        
        # Create table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        output.seek(0)
        filename = f"bus_routes_{datetime.now().strftime('%Y-%m-%d')}.pdf"
        
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid format. Use 'excel' or 'pdf'"
        )


# ==================== STATISTICS ====================

@router.get("/statistics", response_model=StatisticsResponse)
def get_statistics(
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Get dashboard statistics"""
    total_drivers = db.query(Driver).count()
    active_drivers = db.query(Driver).filter(Driver.is_active == True).count()
    total_routes = db.query(BusRoute).count()
    
    # Get active buses from Redis
    active_buses_list = CacheService.get_all_active_buses()
    active_buses = len(active_buses_list)
    
    return StatisticsResponse(
        total_drivers=total_drivers,
        active_drivers=active_drivers,
        total_routes=total_routes,
        active_buses=active_buses,
        last_updated=datetime.utcnow()
    )


# ==================== AUDIT LOG ====================

@router.get("/audit-log", response_model=AuditLogListResponse)
def get_audit_log(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    action_type: Optional[str] = None,
    page: int = 1,
    per_page: int = 50,
    db: Session = Depends(get_db),
    current_admin: Driver = Depends(get_current_admin())
):
    """Get audit log entries with filters"""
    query = db.query(AuditLog).join(Driver, AuditLog.admin_id == Driver.driver_id)
    
    # Apply filters
    if start_date:
        query = query.filter(AuditLog.created_at >= start_date)
    if end_date:
        query = query.filter(AuditLog.created_at <= end_date)
    if action_type and action_type != "ALL":
        query = query.filter(AuditLog.action_type == action_type)
    
    # Get total count
    total = query.count()
    
    # Calculate pagination
    pages = math.ceil(total / per_page)
    offset = (page - 1) * per_page
    
    # Get paginated results
    logs = query.order_by(AuditLog.created_at.desc()).offset(offset).limit(per_page).all()
    
    # Format response
    log_responses = []
    for log in logs:
        admin = db.query(Driver).filter(Driver.driver_id == log.admin_id).first()
        log_responses.append(AuditLogResponse(
            log_id=log.log_id,
            admin_id=log.admin_id,
            admin_name=admin.name if admin else "Unknown",
            action_type=log.action_type,
            entity_type=log.entity_type,
            entity_id=log.entity_id,
            changes=log.changes,
            created_at=log.created_at
        ))
    
    return AuditLogListResponse(
        logs=log_responses,
        total=total,
        page=page,
        pages=pages
    )
